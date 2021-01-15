import openpyxl
from pathlib import Path

weiter = True
verAnzahl = 1
count = 1
listab = []
rows = 50
columns = 7
row = 0
colum = 0
abstand = 62


print("")
print("")
verAnzahl = input("Anzahl der verkäufer: ")
print("")
print("Kopieren und einfügen: ")


wbGes = openpyxl.load_workbook("gesamtliste.xlsx")
sheetGes = wbGes["komplett"]


while(count <= int(verAnzahl)):
    wbVer = openpyxl.load_workbook("Verkäufer" + str(count) + ".xlsx")
    sheetVer = wbVer["Tabelle1"]
    print("Liste: " + str(count))

    for i in range(1, rows + 1):
        listab.append([])

    for r in range(1, rows + 1):
        for c in range(1, columns + 1):
            e = sheetVer.cell(row = r, column = c)
            listab[r - 1].append(e.value)


    for r in range(1, rows + 1):
        for c in range(1, columns + 1):
            sheetGes.cell(row=r+11+(int(count)-1)*abstand, column=c+1).value = listab[r-1][c-1]

    listab.clear()

    count = count + 1

print("speichern...")
wbGes.save("gesamtliste.xlsx")
print("fertig")